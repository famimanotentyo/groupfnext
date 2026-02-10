import openpyxl
import os

file_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(file_path, data_only=True)
    print("=== Sheet Names ===")
    target_sheet = 'テーブル定義書_ ユーザー'
    if target_sheet in wb.sheetnames:
        sheet = wb[target_sheet]
        header_row_index = None
        col_indices = {}
        
        # Search for header row
        for i, row in enumerate(sheet.iter_rows(min_row=1, max_row=20, values_only=True)):
            row_values = [str(cell).strip() if cell else "" for cell in row]
            if '物理名称' in row_values and 'データ型' in row_values:
                header_row_index = i + 1
                print(f"Table Header found at Row {header_row_index}")
                for j, col in enumerate(row):
                    if col:
                        print(f"Index {j}: {col}")
                        col_indices[col] = j
                break
        
        if header_row_index:
             print(f"--- Rows {header_row_index+1}-{header_row_index+15} of {target_sheet} ---")
             for i, row in enumerate(sheet.iter_rows(min_row=header_row_index+1, max_row=header_row_index+15, values_only=True)):
                 print(f"Row {header_row_index+1+i}: {row}")
    else:
        print(f"Sheet '{target_sheet}' not found.")
except ModuleNotFoundError:
    print("Error: openpyxl module not found. Please install it using 'pip install openpyxl'.")
except Exception as e:
    print(f"Error opening Excel file: {e}")
