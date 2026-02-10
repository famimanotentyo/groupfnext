import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    target_sheet_name = 'テーブル一覧②'
    
    if target_sheet_name not in wb.sheetnames:
        print(f"Sheet '{target_sheet_name}' not found. Available sheets: {wb.sheetnames}")
    else:
        sheet = wb[target_sheet_name]
        print(f"--- Searching for headers in {target_sheet_name} ---")
        
        headers = ['No', '論理名称', '物理名称']
        col_map = {}
        
        for r in range(1, 20):
            row_values = []
            for c in range(1, 20):
                cell = sheet.cell(row=r, column=c)
                val = str(cell.value).strip() if cell.value else None
                row_values.append(val)
                
                if val in headers:
                    col_map[val] = (r, c)
            
            print(f"Row {r}: {row_values}")
            
        print("\nFound headers:")
        for k, v in col_map.items():
            print(f"{k}: Row {v[0]}, Col {v[1]}")

except Exception as e:
    print(f"Error: {e}")
