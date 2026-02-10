import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = 'テーブル定義書_ユーザー'
    
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        print(f"--- Inspecting FK rows in {sheet_name} ---")
        
        # Headers (Row 4)
        # 1:No, 3:Logical, 12:Physical, 21:Type, 31:PK, 33:FK, 37:Remarks
        
        for r in range(5, 30):
            fk_val = sheet.cell(row=r, column=33).value # FK column
            if fk_val and str(fk_val).strip() in ['○', 'Yes', 'True']:
                # Found a FK row
                p_name = sheet.cell(row=r, column=12).value
                remarks = sheet.cell(row=r, column=37).value
                print(f"Row {r}: Field='{p_name}', FK='{fk_val}', Remarks='{remarks}'")
                
    else:
        print(f"{sheet_name} not found.")

except Exception as e:
    print(f"Error: {e}")
