import openpyxl
import re

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'
output_puml_path = r'c:\Django課題制作 - コピー\groupf\db_schema.puml'

def main():
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
        
        # 1. Get Table List
        list_sheet = None
        for s in wb.sheetnames:
            if 'テーブル一覧' in s:
                list_sheet = wb[s]
                break
        
        tables = [] # List of dicts: {logical, physical}
        if list_sheet:
            print(f"Reading table list from {list_sheet.title}...")
            col_logical = -1
            col_physical = -1
            header_row = -1
            
            for r in range(1, 10):
                for c in range(1, 20):
                    val = str(list_sheet.cell(row=r, column=c).value).strip() if list_sheet.cell(row=r, column=c).value else ""
                    if val == '論理名称':
                        col_logical = c
                        header_row = r
                    elif val == '物理名称':
                        col_physical = c
                        header_row = r
            
            if col_logical != -1 and col_physical != -1:
                for r in range(header_row + 1, 200):
                    l_val = list_sheet.cell(row=r, column=col_logical).value
                    p_val = list_sheet.cell(row=r, column=col_physical).value
                    if l_val and p_val:
                        tables.append({
                            'logical': str(l_val).strip(),
                            'physical': str(p_val).strip()
                        })
        
        print(f"Found {len(tables)} tables.")
        
        # Map physical -> logical for relationship linking
        phy_to_log = {t['physical']: t['logical'] for t in tables}
        # Also map suffixes to logical for fuzzy matching
        suffix_to_log = {}
        for p, l in phy_to_log.items():
            parts = p.split('_')
            if len(parts) > 1:
                suffix_to_log[parts[-1]] = l
            suffix_to_log[p] = l

        entities = {} # logical_name -> {fields: list}
        
        for t in tables:
            l_name = t['logical']
            # p_name = t['physical']
            
            sheet_name_candidates = [f"テーブル定義書_{l_name}", f"テーブル設計書_{l_name}"]
            target_sheet = None
            
            for cand in sheet_name_candidates:
                if cand in wb.sheetnames:
                    target_sheet = wb[cand]
                    break
            
            if not target_sheet:
                # Try fuzzy search
                for s in wb.sheetnames:
                    if l_name in s and 'テーブル' in s:
                        target_sheet = wb[s]
                        break
            
            if not target_sheet:
                print(f"Sheet for {l_name} not found. Skipping.")
                continue
                
            # Headers
            col_l_field = -1 # Logical Field Name
            col_p_field = -1
            col_type = -1
            col_pk = -1
            col_fk = -1
            header_row_idx = -1
            
            for r in range(1, 10):
                for c in range(1, 50):
                    val = str(target_sheet.cell(row=r, column=c).value).strip() if target_sheet.cell(row=r, column=c).value else ""
                    if val == '論理名称':
                        col_l_field = c
                        header_row_idx = r
                    elif val == '物理名称':
                        col_p_field = c
                    elif val == 'データ型':
                        col_type = c
                    elif val == 'PK':
                        col_pk = c
                    elif val == 'FK':
                        col_fk = c

            if col_l_field == -1:
                # Fallback based on typical inspection
                col_l_field = 3
                col_p_field = 12 # Approximation
                col_type = 21
                header_row_idx = 4
            
            fields = []
            for r in range(header_row_idx + 1, 100):
                f_log = target_sheet.cell(row=r, column=col_l_field).value
                f_phys = target_sheet.cell(row=r, column=col_p_field).value if col_p_field != -1 else ""
                
                if not f_log and not f_phys:
                    break
                
                f_type = target_sheet.cell(row=r, column=col_type).value if col_type != -1 else ""
                f_pk = target_sheet.cell(row=r, column=col_pk).value if col_pk != -1 else ""
                f_fk = target_sheet.cell(row=r, column=col_fk).value if col_fk != -1 else ""
                
                is_pk = str(f_pk).strip() in ['○', 'Yes', 'True', 'PK']
                is_fk = str(f_fk).strip() in ['○', 'Yes', 'True', 'FK']
                
                fields.append({
                    'logical': str(f_log).strip() if f_log else "",
                    'physical': str(f_phys).strip() if f_phys else "",
                    'type': str(f_type).strip(),
                    'pk': is_pk,
                    'fk': is_fk
                })
                
            entities[l_name] = {
                'physical': t['physical'],
                'fields': fields
            }

        # 3. Generate PlantUML
        lines = []
        lines.append("@startuml")
        lines.append("!theme plain")
        lines.append("hide circle")
        lines.append("skinparam linetype ortho")
        lines.append("")
        
        # Classes
        for l_name, data in entities.items():
            lines.append(f'class "{l_name}" {{')
            for f in data['fields']:
                pk_mark = "" # Reference didn't use explicit PK mark in field name, just type or logic. 
                # Reference: "- 顧客ID: int".
                # Let's use Fk type if it is FK.
                
                ftype = f['type']
                if f['fk']:
                    ftype = "Fk"
                
                fname = f['logical']
                if not fname: fname = f['physical']
                
                # Visibility '-' as per reference
                lines.append(f'  - {fname}: {ftype}')
                
            lines.append("}")
            lines.append("")

        # Relationships
        relationships = set()
        
        # Special manual mappings for Logical Names hard to guess
        special_map_log = {
            '作成者': 'ユーザー',
            '更新者': 'ユーザー',
            '承認者': 'ユーザー',
            '依頼者': 'ユーザー',
            '担当者': 'ユーザー',
            '部下': 'ユーザー',
            '上司': 'ユーザー',
            '担当上司': 'ユーザー',
            '対象部下': 'ユーザー',
            '質問者': 'ユーザー',
            '回答者': 'ユーザー',
            '差出人': 'ユーザー',
            '受信者': 'ユーザー',
            '通知先': 'ユーザー',
            '送信者': 'ユーザー',
            '閲覧ユーザー': 'ユーザー',
            '関連部署': '部署',
            '所属グループ': '部署',
            '権限': '役割マスタ',
            'タスク状態': 'タスク状態マスタ',
            '元チャット': '相談',
            '種類': 'スケジュール種別マスタ',
            # M2M Targets
            'タグ': 'タグ',
            '関連タグ': 'タグ',
            '取り掛かり中の人': 'ユーザー',
            '完了とした人': 'ユーザー',
            '参加者': 'ユーザー',
            'ブックマーク': 'ユーザー',
            'グループ': '部署', # Assuming group maps to deployment/role or skip
        }

        for l_name, data in entities.items():
            for f in data['fields']:
                is_m2m = f['type'] == 'ManyToManyField'
                
                if f['fk'] or is_m2m:
                    fname = f['logical'] # e.g. "所属グループ"
                    phys_fname = f['physical'] # e.g. "department_id"
                    
                    target_log = None
                    
                    # 1. Try Map with Logical Name
                    if fname in special_map_log:
                        target_log = special_map_log[fname]
                    
                    # 2. Try suffix match on Physical Name (only if not M2M, usually)
                    if not target_log and phys_fname and not is_m2m:
                        base = phys_fname.lower()
                        if base.endswith('_id'): base = base[:-3]
                        
                        if base in suffix_to_log:
                            target_log = suffix_to_log[base]
                            
                    # 3. Fuzzy match logical name to other table logical names
                    if not target_log:
                        candidates = []
                        current_app = data['physical'].split('_')[0] if '_' in data['physical'] else ""
                        
                        for cand_l, cand_data in entities.items():
                            if cand_l == l_name: continue
                            
                            if fname in cand_l or cand_l in fname:
                                cand_app = cand_data['physical'].split('_')[0] if '_' in cand_data['physical'] else ""
                                candidates.append((cand_l, cand_app))
                        
                        same_app = [c[0] for c in candidates if c[1] == current_app]
                        if same_app:
                            target_log = same_app[0]
                        elif candidates:
                            target_log = candidates[0][0]
                            
                    if target_log and target_log in entities:
                        if target_log != l_name:
                            if is_m2m:
                                # Many to Many
                                rel = f'"{target_log}" "0..*" -- "0..*" "{l_name}"'
                            else:
                                # One to Many
                                rel = f'"{target_log}" "1" -- "0..*" "{l_name}"'
                            relationships.add(rel)

        lines.extend(sorted(list(relationships)))
        lines.append("@enduml")
        
        with open(output_puml_path, 'w', encoding='utf-8') as f:
            f.write("\n".join(lines))
            
        print(f"Generated Logically-named PlantUML at {output_puml_path}")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
