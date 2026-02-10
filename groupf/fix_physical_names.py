import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

def main():
    try:
        wb = openpyxl.load_workbook(excel_path)
        
        # 1. Identify Source List Sheet
        print(f"Available sheets: {wb.sheetnames}")
        
        source_sheet_name = None
        # Try exact match first
        if 'テーブル一覧①' in wb.sheetnames:
            source_sheet_name = 'テーブル一覧①'
        elif 'テーブル一覧②' in wb.sheetnames:
            source_sheet_name = 'テーブル一覧②'
        else:
            # Try fuzzy match
            for s in wb.sheetnames:
                if 'テーブル一覧' in s:
                    source_sheet_name = s
                    break
        
        if not source_sheet_name:
            print("Error: Could not find any sheet matching 'テーブル一覧'.")
            return

        print(f"Using source sheet: {source_sheet_name}")
        source_sheet = wb[source_sheet_name]

        # 2. Map Logical Map -> Physical Name
        # Find headers to be robust
        col_logical = None
        col_physical = None
        
        for r in range(1, 10):
            for c in range(1, 20):
                val = str(source_sheet.cell(row=r, column=c).value).strip() if source_sheet.cell(row=r, column=c).value else ""
                if val == '論理名称':
                    col_logical = c
                elif val == '物理名称':
                    col_physical = c
            if col_logical and col_physical:
                break
        
        if not (col_logical and col_physical):
            print("Could not find '論理名称' or '物理名称' columns in source sheet.")
            # Fallback to observed values if needed?
            # heuristic: Logical=3, Physical=14
            print("Falling back to expected columns: Logical=3, Physical=14")
            col_logical = 3
            col_physical = 14
        
        name_map = {}
        # Iterate rows
        # Assuming data starts after header. Let's start scanning from row 5 down.
        for r in range(5, 200):
            l_name = source_sheet.cell(row=r, column=col_logical).value
            p_name = source_sheet.cell(row=r, column=col_physical).value
            
            if l_name:
                name_map[str(l_name).strip()] = str(p_name).strip() if p_name else ""
            else:
                # Stop if consecutive empty logical names?
                # or just continue a bit
                continue
                
        print(f"Loaded {len(name_map)} mappings from {source_sheet_name}.")
        
        # 3. Update Definition Sheets
        updated_count = 0
        for sheet_name in wb.sheetnames:
            if sheet_name.startswith('テーブル定義書_') or sheet_name.startswith('テーブル設計書_'):
                # Extract logical name from sheet name?
                # The sheet name format is "テーブル定義書_LogicalName"
                # But sometimes it might differ slightly.
                # However, the user said "referencing Table List 1".
                # The Table List 1 has Logical Name.
                # The TABLE DEFINITION SHEET also has Logical Name in header (Row 1, Col 29... wait, check inspection)
                # Inspection: Row 1, Col 23: 論理名称, Row 1, Col 29: ユーザー (Logical Name)
                # Inspection: Row 2, Col 23: 物理名称, Row 2, Col 29: accounts_user (Physical Name TARGET)
                
                sheet = wb[sheet_name]
                
                # Read Logical Name from sheet header to be sure
                # It seems to be at Row 1, Column 29 based on inspection of 'テーブル定義書_ユーザー'
                # Let's verify loc of "論理名称" label first
                
                target_logical_val = sheet.cell(row=1, column=29).value
                
                # If that's empty or doesn't match, maybe finding "論理名称" label is safer
                # Inspection said: Row 1, Col 23 is "論理名称".
                # So value should be at Col 29 (merged? or just offset).
                # Actually, inspection showed "Row 1, Col 29: ユーザー".
                
                # Let's try to match by sheet name suffix if header read fails or just as primary key
                # Sheet name: テーブル定義書_ユーザー -> Logical: ユーザー
                # This aligns with the map keys.
                
                # Parsing sheet name
                parts = sheet_name.split('_', 1)
                if len(parts) > 1:
                    sheet_logical_name = parts[1]
                else:
                    sheet_logical_name = target_logical_val
                
                if sheet_logical_name in name_map:
                    correct_physical = name_map[sheet_logical_name]
                    
                    # Target Cell for Physical Name
                    # Inspection: Row 2, Col 23 is "物理名称".
                    # Inspection: Row 2, Col 29 is the value.
                    
                    # Update cell Row 2, Col 29
                    current_val = sheet.cell(row=2, column=29).value
                    if current_val != correct_physical:
                        print(f"Updating {sheet_name}: {current_val} -> {correct_physical}")
                        sheet.cell(row=2, column=29).value = correct_physical
                        updated_count += 1
                    else:
                        # print(f"Skipping {sheet_name}: already correct ({current_val})")
                        pass
                else:
                    print(f"Warning: No mapping found for sheet '{sheet_name}' (Logical: {sheet_logical_name})")

        wb.save(excel_path)
        print(f"Update complete. {updated_count} sheets updated.")

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()

if __name__ == '__main__':
    main()
