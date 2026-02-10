import openpyxl
import json
import re
import os

# Paths
json_path = r'c:\Django課題制作 - コピー\groupf\models_schema.json'
excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

def load_json(path):
    encodings = ['utf-8', 'utf-16', 'cp932']
    for enc in encodings:
        try:
            with open(path, 'r', encoding=enc) as f:
                return json.load(f)
        except:
            continue
    return None

def main():
    # 1. Load JSON
    models_data = load_json(json_path)
    if not models_data:
        print("Failed to load models_schema.json")
        exit(1)
        
    # Map verbose_name -> {app, model}
    # Note: verbose_name might not be unique globally, but assuming it matches sheet names
    model_map = {}
    for entry in models_data:
        v_name = entry['verbose_name']
        model_map[v_name] = {
            'app': entry['app'],
            'model': entry['model']
        }
    
    print(f"Loaded {len(model_map)} models from JSON.")

    # 2. Load Excel
    try:
        wb = openpyxl.load_workbook(excel_path)
    except Exception as e:
        print(f"Failed to load Excel: {e}")
        exit(1)
        
    target_sheet_name = 'テーブル一覧②'
    if target_sheet_name not in wb.sheetnames:
        print(f"Sheet '{target_sheet_name}' not found.")
        exit(1)
    
    sheet = wb[target_sheet_name]
    
    # 3. Find Headers
    header_row = None
    col_no = None
    col_logical = None
    col_physical = None
    
    # Search first 20 rows
    for r in range(1, 21):
        for c in range(1, 21):
            val = sheet.cell(row=r, column=c).value
            if val:
                s_val = str(val).strip()
                if s_val == 'No':
                    col_no = c
                    header_row = r
                elif s_val == '論理名称':
                    col_logical = c
                    header_row = r
                elif s_val == '物理名称':
                    col_physical = c
                    header_row = r
        
        if col_no and col_logical and col_physical:
            break
            
    if not (col_no and col_logical and col_physical):
        print("Could not find all headers (No, 論理名称, 物理名称) in the first 20 rows.")
        # Fallback inspection print
        if header_row:
             print(f"Found headers partially at row {header_row}: No={col_no}, Logical={col_logical}, Physical={col_physical}")
        exit(1)
        
    print(f"Headers found at Row {header_row}: No={col_no}, Logical={col_logical}, Physical={col_physical}")
    
    # 4. Gather Sheets and Data
    # Identify sheets starting with 'テーブル定義書_' (or 'テーブル設計書_' as per prompt example, checking both)
    # The generation script uses 'テーブル定義書_', but prompt says 'テーブル設計書_'. I will check for 'テーブル定義書_' since that's what generate_all_sheets.py produces.
    
    data_rows = []
    
    # Sort sheet names to match tab order? Or just iterate?
    # Usually list order logic is preferred. Let's filter and organize.
    
    # We want to exclude the template sheet potentially?
    # generate_all_sheets uses 'テーブル定義書_ ユーザー' as template.
    # But usually we list all the generated tables.
    
    # Regex to capture the part after the prefix
    # Matches 'テーブル定義書_Something' or 'テーブル設計書_Something'
    pattern = re.compile(r'^(?:テーブル定義書|テーブル設計書)[_＿](.+)$')
    
    for sheet_name in wb.sheetnames:
        match = pattern.match(sheet_name)
        if match:
            logical_name_from_sheet = match.group(1)
            
            # Skip the template itself if it's purely a template (often checked by name)
            # But here 'テーブル定義書_ ユーザー' is arguably a real table too (User model).
            # The prompt example: "Sheet 'テーブル設計書＿ユーザー' -> Logical Name 'ユーザー'"
            
            # Check mapping
            if logical_name_from_sheet in model_map:
                info = model_map[logical_name_from_sheet]
                physical_name = f"{info['app']}_{info['model']}".lower() # e.g. accounts_user
            else:
                # If not found in JSON (maybe manually added or name mismatch)
                print(f"Warning: Sheet '{sheet_name}' (Logical: {logical_name_from_sheet}) not found in JSON map.")
                physical_name = "-"
            
            data_rows.append({
                'logical': logical_name_from_sheet,
                'physical': physical_name
            })
            
    print(f"Found {len(data_rows)} sheets to list.")
    
    # 5. Write to Sheet
    # Start writing from header_row + 1
    start_row = header_row + 1
    
    # Clear existing data first?
    # Look for the end of the list. We'll simply overwrite and clear remaining if it was longer before.
    # But to be safe, let's just write and clear afterwards.
    
    for i, data in enumerate(data_rows):
        current_row = start_row + i
        
        # Write No
        sheet.cell(row=current_row, column=col_no).value = i + 1
        
        # Write Logical Name
        sheet.cell(row=current_row, column=col_logical).value = data['logical']
        
        # Write Physical Name
        sheet.cell(row=current_row, column=col_physical).value = data['physical']
        
    # Clear remaining rows if previous list was longer
    # Heuristic: check if next row has a number in 'No' column
    row_to_clean = start_row + len(data_rows)
    while True:
        cell_no = sheet.cell(row=row_to_clean, column=col_no)
        if cell_no.value is None:
            break
        
        # Clear row
        sheet.cell(row=row_to_clean, column=col_no).value = None
        sheet.cell(row=row_to_clean, column=col_logical).value = None
        sheet.cell(row=row_to_clean, column=col_physical).value = None
        
        row_to_clean += 1
        if row_to_clean > 1000: # Safety break
            break
            
    wb.save(excel_path)
    print("Successfully updated table list.")

if __name__ == '__main__':
    main()
