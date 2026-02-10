import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    sheet_name = 'テーブル定義書_タスク'
    
    if sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        # Physical name label at Row 2, Col 23
        # Value at Row 2, Col 29
        
        label = sheet.cell(row=2, column=23).value
        value = sheet.cell(row=2, column=29).value
        
        print(f"Sheet: {sheet_name}")
        print(f"Label (R2C23): {label}")
        print(f"Value (R2C29): {value}")
        
    else:
        print(f"Sheet {sheet_name} not found.")

except Exception as e:
    print(f"Error: {e}")
