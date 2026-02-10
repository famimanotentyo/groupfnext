import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    
    # 1. Inspect Table List 1
    list_sheet_name = 'テーブル一覧①'
    if list_sheet_name in wb.sheetnames:
        sheet = wb[list_sheet_name]
        print(f"--- {list_sheet_name} (First 10 rows) ---")
        for r in range(1, 11):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 10)]
            print(f"Row {r}: {row_vals}")
    else:
        print(f"Sheet '{list_sheet_name}' not found.")

    # 2. Inspect a Definition Sheet
    def_sheet_name = 'テーブル定義書_ユーザー'
    if def_sheet_name not in wb.sheetnames:
        # Try to find another one
        candidates = [s for s in wb.sheetnames if s.startswith('テーブル定義書_')]
        if candidates:
            def_sheet_name = candidates[0]
    
    if def_sheet_name in wb.sheetnames:
        sheet = wb[def_sheet_name]
        print(f"\n--- {def_sheet_name} (Header Area Rows 1-5) ---")
        for r in range(1, 6):
            row_vals = [sheet.cell(row=r, column=c).value for c in range(1, 40)] # Check wide range
            print(f"Row {r}: {row_vals}")
            
    else:
        print("No definition sheet found.")

except Exception as e:
    print(f"Error: {e}")
