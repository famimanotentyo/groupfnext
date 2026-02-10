import openpyxl
import shutil
import os

# File paths
original_file = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'
backup_file = r'c:\Django課題制作 - コピー\26_テーブル設計書_backup.xlsx'

# Create backup
if not os.path.exists(backup_file):
    shutil.copy2(original_file, backup_file)
    print(f"Backup created: {backup_file}")
else:
    print(f"Backup already exists: {backup_file}")

# Data to write
schema_data = [
    {"logical_name": "ID", "physical_name": "id", "type": "BigAuto", "length": "-", "default": "-", "pk": "○", "fk": "", "nn": "○", "remarks": "自動連番"},
    {"logical_name": "パスワード", "physical_name": "password", "type": "Varchar", "length": 128, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": "AbstractBaseUser"},
    {"logical_name": "最終ログイン日時", "physical_name": "last_login", "type": "DateTime", "length": "-", "default": "-", "pk": "", "fk": "", "nn": "", "remarks": "AbstractBaseUser"},
    {"logical_name": "スーパーユーザー", "physical_name": "is_superuser", "type": "Boolean", "length": "-", "default": "False", "pk": "", "fk": "", "nn": "○", "remarks": "PermissionsMixin"},
    {"logical_name": "社員番号", "physical_name": "employee_number", "type": "Varchar", "length": 20, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": "Unique"},
    {"logical_name": "姓", "physical_name": "last_name", "type": "Varchar", "length": 50, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "名", "physical_name": "first_name", "type": "Varchar", "length": 50, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "セイ", "physical_name": "last_name_kana", "type": "Varchar", "length": 50, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "メイ", "physical_name": "first_name_kana", "type": "Varchar", "length": 50, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "メールアドレス", "physical_name": "email", "type": "Email", "length": 254, "default": "-", "pk": "", "fk": "", "nn": "○", "remarks": "Unique"},
    {"logical_name": "電話番号", "physical_name": "phone_number", "type": "Varchar", "length": 15, "default": "-", "pk": "", "fk": "", "nn": "", "remarks": "Null可"},
    {"logical_name": "アイコン画像", "physical_name": "avatar", "type": "Image", "length": "-", "default": "-", "pk": "", "fk": "", "nn": "", "remarks": "Null可"},
    {"logical_name": "所属グループ", "physical_name": "department_id", "type": "BigInt", "length": "-", "default": "-", "pk": "", "fk": "○", "nn": "", "remarks": "Null可"},
    {"logical_name": "権限", "physical_name": "role_id", "type": "BigInt", "length": "-", "default": "-", "pk": "", "fk": "○", "nn": "", "remarks": "Null可"},
    {"logical_name": "生年月日", "physical_name": "birth_date", "type": "Date", "length": "-", "default": "-", "pk": "", "fk": "", "nn": "", "remarks": "Null可"},
    {"logical_name": "入社日", "physical_name": "hire_date", "type": "Date", "length": "-", "default": "-", "pk": "", "fk": "", "nn": "", "remarks": "Null可"},
    {"logical_name": "登録日", "physical_name": "date_joined", "type": "DateTime", "length": "-", "default": "timezone.now", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "初回設定完了", "physical_name": "is_initial_setup_completed", "type": "Boolean", "length": "-", "default": "False", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "仮パスワード有効期限", "physical_name": "temp_password_expires_at", "type": "DateTime", "length": "-", "default": "-", "pk": "", "fk": "", "nn": "", "remarks": "Null可"},
    {"logical_name": "有効", "physical_name": "is_active", "type": "Boolean", "length": "-", "default": "True", "pk": "", "fk": "", "nn": "○", "remarks": ""},
    {"logical_name": "管理サイトアクセス権限", "physical_name": "is_staff", "type": "Boolean", "length": "-", "default": "False", "pk": "", "fk": "", "nn": "○", "remarks": ""},
]

try:
    wb = openpyxl.load_workbook(original_file)
    target_sheet_name = 'テーブル定義書_ ユーザー'
    
    if target_sheet_name not in wb.sheetnames:
        print(f"Error: Sheet '{target_sheet_name}' not found.")
        exit(1)
        
    sheet = wb[target_sheet_name]
    
    # Identify header row (around row 4)
    header_row_index = 4
    col_map = {}
    
    # Map headers to column indices (1-based)
    for cell in sheet[header_row_index]:
        if cell.value:
            val = str(cell.value).strip()
            if val == 'No': col_map['no'] = cell.column
            elif val == '論理名称': col_map['logical_name'] = cell.column
            elif val == '物理名称': col_map['physical_name'] = cell.column
            elif val == 'データ型': col_map['type'] = cell.column
            elif val == '桁数': col_map['length'] = cell.column
            elif val == '初期値': col_map['default'] = cell.column
            elif val == 'PK': col_map['pk'] = cell.column
            elif val == 'FK': col_map['fk'] = cell.column
            elif val == 'NN': col_map['nn'] = cell.column # Assuming NN or Not Null
            elif val == '備考': col_map['remarks'] = cell.column

    # If mapping failed for some reason, use fallback indices derived from analysis
    # No:1, Logical:3, Physical:12, Type:21, Length:26, Default:28, PK:31, FK:35, NN:37, Remarks:42
    if 'physical_name' not in col_map:
        print("Warning: Could not auto-detect columns. Using hardcoded indices.")
        col_map = {
            'no': 1,
            'logical_name': 3,
            'physical_name': 12,
            'type': 21,
            'length': 26,
            'default': 28,
            'pk': 31,
            'fk': 35,
            'nn': 37,
            'remarks': 42
        }
    
    print(f"Column Mapping: {col_map}")

    # Write data starting from row 5
    start_row = 5
    for i, data in enumerate(schema_data):
        row_num = start_row + i
        
        # Write No
        if 'no' in col_map:
            sheet.cell(row=row_num, column=col_map['no']).value = i + 1
            
        # Write other fields
        for key, val in data.items():
            if key in col_map:
                sheet.cell(row=row_num, column=col_map[key]).value = val
                
    wb.save(original_file)
    print("Excel updated successfully.")

except Exception as e:
    print(f"Error: {e}")
