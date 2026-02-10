import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    target_sheet_name = 'テーブル一覧②'
    
    if target_sheet_name in wb.sheetnames:
        sheet = wb[target_sheet_name]
        print(f"--- Data in {target_sheet_name} (Rows 5-15) ---")
        # Assuming header is at row 4 as found by update script
        for r in range(5, 16):
            row_values = []
            for c in range(1, 15): # Logical is 3, Physical is 14
                 val = sheet.cell(row=r, column=c).value
                 row_values.append(val)
            print(f"Row {r}: {row_values}")
    else:
        print(f"Sheet {target_sheet_name} not found.")

    print("\n--- Sheet Names starting with テーブル定義書 ---")
    for name in wb.sheetnames:
        if name.startswith('テーブル定義書'):
            print(name)

except Exception as e:
    print(f"Error: {e}")
