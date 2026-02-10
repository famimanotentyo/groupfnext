import json
import openpyxl
import os

json_path = r'c:\Django課題制作 - コピー\groupf\models_schema.json'
excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

with open(json_path, 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = openpyxl.load_workbook(excel_path, read_only=True)
sheet_names = wb.sheetnames

print(f"Total models in JSON: {len(data)}")
print(f"Total sheets in Excel: {len(sheet_names)}")

generated_sheets = [s for s in sheet_names if s.startswith('テーブル定義書_')]
print(f"Generated sheets: {len(generated_sheets)}")

missing = []
for model in data:
    verbose_name = model['verbose_name']
    expected_sheet = f"テーブル定義書_{verbose_name}"[:31]
    # Simple replacement matching generate_all_sheets.py
    import re
    expected_sheet = re.sub(r'[\\/*?:\[\]]', '_', expected_sheet)
    
    if expected_sheet not in sheet_names:
        missing.append(expected_sheet)

if missing:
    print(f"Missing sheets ({len(missing)}):")
    for m in missing:
        print(f" - {m}")
else:
    print("All expected sheets are present.")

# Check content of one sheet (e.g., 'テーブル定義書_ユーザー')
target = 'テーブル定義書_ユーザー'
if target in sheet_names:
    sheet = wb[target]
    print(f"\nVerification of content for {target}:")
    # Header check
    print(f" Row 1, Col 29: {sheet.cell(row=1, column=29).value}")
    print(f" Row 5, Col 3: {sheet.cell(row=5, column=3).value} (Expected: ID or 論理名)")
