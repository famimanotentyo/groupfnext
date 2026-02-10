import openpyxl
import json
import shutil
import os
import re

# Paths
json_path = r'c:\Django課題制作 - コピー\groupf\models_schema.json'
excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'
backup_path = r'c:\Django課題制作 - コピー\26_テーブル設計書_backup_gen.xlsx'

# Load JSON with encoding detection fallback
models_data = None
encodings = ['utf-8', 'utf-16', 'cp932']
for enc in encodings:
    try:
        with open(json_path, 'r', encoding=enc) as f:
            models_data = json.load(f)
        print(f"Successfully loaded JSON with encoding: {enc}")
        break
    except Exception as e:
        print(f"Failed to load JSON with {enc}: {e}")

if models_data is None:
    print("Could not load JSON file with tried encodings.")
    exit(1)

# Backup Excel
try:
    if not os.path.exists(backup_path):
        shutil.copy2(excel_path, backup_path)
        print(f"Backup created: {backup_path}")
    else:
        print(f"Backup already exists: {backup_path}")
except Exception as e:
    print(f"Error creating backup: {e}")
    exit(1)

try:
    wb = openpyxl.load_workbook(excel_path)
    
    # Determined template sheet
    template_sheet_name = 'テーブル定義書_ ユーザー'
    if template_sheet_name not in wb.sheetnames:
        print(f"Template sheet '{template_sheet_name}' not found.")
        # Try to find any sheet starting with テーブル定義書
        candidates = [s for s in wb.sheetnames if s.startswith('テーブル定義書')]
        if candidates:
            template_sheet_name = candidates[0]
            print(f"Using '{template_sheet_name}' as template.")
        else:
            print("No suitable template sheet found.")
            exit(1)
            
    template_sheet = wb[template_sheet_name]
    
    col_map = {} 
    
    # Find Header Row for Columns
    header_row_index = 4
    for cell in template_sheet[header_row_index]:
        if cell.value:
            val = str(cell.value).strip()
            if val == 'No': col_map['no'] = cell.column
            elif val == '論理名称': col_map['logical_name'] = cell.column
            elif val == '物理名称': col_map['physical_name'] = cell.column
            elif val == 'データ型': col_map['type'] = cell.column
            elif val == '桁数': col_map['length'] = cell.column
            elif val == '初期値': col_map['default'] = cell.column
            elif val == 'PK': col_map['pk'] = cell.column
            elif val == 'FK': col_map['fk'] = cell.column
            elif val == 'NN': col_map['nn'] = cell.column
            elif val == '備考': col_map['remarks'] = cell.column
            
    if not col_map: 
        print("Failed to map columns. Check header row.")
        exit(1)

    # Process each model
    for model in models_data:
        app_label = model['app']
        model_name = model['model']
        verbose_name = model['verbose_name']
        
        # Sheet Name
        new_sheet_name = f"テーブル定義書_{verbose_name}"
        
        # Excel sheet name limit is 31 chars
        # prohibited chars: : \ / ? * [ ]
        safe_name = re.sub(r'[\\/*?:\[\]]', '_', new_sheet_name)[:31]
        
        if safe_name in wb.sheetnames:
            print(f"Sheet '{safe_name}' already exists. Overwriting/Updating...")
            target_sheet = wb[safe_name]
            # Clear existing data rows (from row 5 downwards)
            max_row = target_sheet.max_row
            if max_row < 5: max_row = 5
            # We want to clear content but keep formatting if possible.
            # Ideally delete rows and copy styles again, but simple clear is safer.
            for r in range(5, max_row + 20): # +20 buffer
                for c in range(1, 50):
                     target_sheet.cell(row=r, column=c).value = None
        else:
            print(f"Creating sheet '{safe_name}'...")
            target_sheet = wb.copy_worksheet(template_sheet)
            target_sheet.title = safe_name
            
        # Update Metadata in Row 1 if possible
        # Heuristic: Search Row 1 for 'ユーザー' or logical name placeholder and replace
        updated_meta = False
        for c in range(1, 50):
            val = target_sheet.cell(row=1, column=c).value
            if val and (str(val).strip() == 'ユーザー'):
                target_sheet.cell(row=1, column=c).value = verbose_name
                updated_meta = True
                break
        
        if not updated_meta:
            # Fallback: exact coordinate from previous observation
            # Row 1, Index 28 -> Column 29
            target_sheet.cell(row=1, column=29).value = verbose_name

        # Write Fields
        fields = model['fields']
        start_row = 5
        for i, field in enumerate(fields):
            current_row = start_row + i
            
            # No
            if 'no' in col_map: target_sheet.cell(row=current_row, column=col_map['no']).value = i + 1
            # Logical
            if 'logical_name' in col_map: target_sheet.cell(row=current_row, column=col_map['logical_name']).value = field['logical_name']
            # Physical
            if 'physical_name' in col_map: target_sheet.cell(row=current_row, column=col_map['physical_name']).value = field['physical_name']
            # Type
            if 'type' in col_map: target_sheet.cell(row=current_row, column=col_map['type']).value = field['type']
            # Length
            if 'length' in col_map: target_sheet.cell(row=current_row, column=col_map['length']).value = field['length']
            # Default
            if 'default' in col_map: target_sheet.cell(row=current_row, column=col_map['default']).value = field['default']
            # PK
            if 'pk' in col_map: target_sheet.cell(row=current_row, column=col_map['pk']).value = field['pk']
            # FK
            if 'fk' in col_map: target_sheet.cell(row=current_row, column=col_map['fk']).value = field['fk']
            # NN
            if 'nn' in col_map: target_sheet.cell(row=current_row, column=col_map['nn']).value = field['nn']
            # Remarks
            if 'remarks' in col_map: target_sheet.cell(row=current_row, column=col_map['remarks']).value = field['remarks']

    wb.save(excel_path)
    print("All sheets generated successfully.")

except Exception as e:
    print(f"Error processing Excel: {e}")
    import traceback
    traceback.print_exc()
