import openpyxl

excel_path = r'c:\Django課題制作 - コピー\26_テーブル設計書.xlsx'

try:
    wb = openpyxl.load_workbook(excel_path)
    print("--- Sheet Names ---")
    for name in wb.sheetnames:
        print(name)
        
    print("\n--- Inspecting Header of テーブル定義書_ユーザー ---")
    if 'テーブル定義書_ユーザー' in wb.sheetnames:
        sheet = wb['テーブル定義書_ユーザー']
        found = False
        for r in range(1, 10):
            for c in range(1, 40): # Check a reasonable range
                val = sheet.cell(row=r, column=c).value
                if val:
                    # Print everything for deep inspection
                    print(f"Row {r}, Col {c}: {val}")
                    if '物理名称' in str(val):
                        print(f"FOUND '物理名称' at Row {r}, Col {c}")
                        # Check subsequent columns to see where the value should go
                        # Usually it's in the next cell or a merged cell.
                        # Let's inspect +5 columns
                        for offset in range(1, 6):
                             v_off = sheet.cell(row=r, column=c+offset).value
                             print(f"  +Offset {offset} (Col {c+offset}): {v_off}")
                        found = True
        
        if not found:
            print("Could not find '物理名称' string in first 10 rows.")
    else:
        print("Sheet 'テーブル定義書_ユーザー' not found.")

except Exception as e:
    print(f"Error: {e}")
